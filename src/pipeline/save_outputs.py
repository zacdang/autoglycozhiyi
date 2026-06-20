"""
Post-processing & Provenance — Module 7.

Two entry points:

post_process_and_save()  ← NEW primary path
    Takes the scheme fill_results (from fill_scheme_missing_fields()) and runs
    a final GPT-4o standardization pass using the prompt from
    configs/prompts/06_Post-processing & Provenance.md.
    Saves: final scheme JSON, intermediate files, provenance summary.

save_outputs()  ← legacy path (kept for backward compatibility)
    Takes the flat ReactionRecord and saves it as before.

Saved files per paper:
  data/outputs/{paper_id}_schemes.json          ← final scheme JSON
  data/outputs/{paper_id}_provenance.json       ← provenance summary
  data/outputs/{paper_id}_extraction_log.json   ← module-level audit log
  data/outputs/{paper_id}_solution.csv          ← solution-phase rows
  data/outputs/{paper_id}_solid.csv             ← solid-phase rows
  data/outputs/test.xlsx                        ← combined Excel workbook
  data/outputs/{paper_id}_unresolved_fields.csv ← fields marked NR
  data/intermediate/{paper_id}_merged.json
  data/intermediate/{paper_id}_unresolved_ids.json
  data/intermediate/{paper_id}_report.json
"""

import csv
import json
from pathlib import Path
from typing import Optional, List

from configs import settings
from src.models.reaction_record import ReactionRecord
from src.utils.json_utils import save_json
from src.utils.logging_utils import get_logger

logger = get_logger(__name__)

_PROMPT_PATH = (
    Path(__file__).parents[2]
    / "configs" / "prompts"
    / "06_Post-processing & Provenance.md"
)


# ── NEW primary path ──────────────────────────────────────────────────────────

def post_process_and_save(
    paper_id: str,
    fill_results: List[dict],
    scheme_extractions: List[dict],
    completeness_reports: List[dict],
    unified: dict,
    id_dict: dict,
    output_dir: Optional[Path] = None,
    intermediate_dir: Optional[Path] = None,
    doi: str = "",
    si_data: Optional[dict] = None,
) -> dict:
    """
    Run Module 7 post-processing on each filled scheme, then save all outputs.

    Parameters
    ----------
    paper_id             : Paper identifier.
    fill_results         : Output of fill_scheme_missing_fields() — one per scheme.
    scheme_extractions   : Output of run_figure_extraction().
    completeness_reports : Output of check_completeness().
    unified              : Merged extraction dict (for intermediate save).
    id_dict              : Identifier dictionary (for unresolved save).
    output_dir           : Where to save final outputs.
    intermediate_dir     : Where to save debug files.

    Returns
    -------
    dict mapping output_name → str(path)
    """
    output_dir       = Path(output_dir       or settings.OUTPUT_DIR)
    intermediate_dir = Path(intermediate_dir or settings.INTERMEDIATE_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    intermediate_dir.mkdir(parents=True, exist_ok=True)

    prompt_text = _load_prompt()
    paths       = {}

    # ── Post-process each scheme ───────────────────────────────────────────────
    final_schemes = []
    provenance_summaries = []

    for fill_result in fill_results:
        figure_id = fill_result.get("figure_id", "unknown")
        logger.info(f"[post_process_and_save] Post-processing {figure_id} …")

        if settings.PIPELINE_MODE == "real" and prompt_text:
            pp_result = _postprocess_with_llm(fill_result, prompt_text, figure_id)
        else:
            pp_result = _postprocess_rule_based(fill_result, figure_id)

        final_schemes.append({
            "figure_id":        figure_id,
            "final_output":     pp_result.get("final_output", {}),
            "unresolved_items": pp_result.get("unresolved_items", []),
        })
        provenance_summaries.append({
            "figure_id":        figure_id,
            "provenance_summary": pp_result.get("provenance_summary", {}),
        })

    # ── Save outputs ───────────────────────────────────────────────────────────

    # 1. Final scheme extraction (primary output)
    schemes_path = output_dir / f"{paper_id}_schemes.json"
    save_json({"paper_id": paper_id, "schemes": final_schemes}, schemes_path)
    paths["final_schemes"] = str(schemes_path)

    # 2. Provenance summary
    prov_path = output_dir / f"{paper_id}_provenance.json"
    save_json({"paper_id": paper_id, "provenance": provenance_summaries}, prov_path)
    paths["provenance"] = str(prov_path)

    # 3. Intermediate merged extraction
    merged_path = intermediate_dir / f"{paper_id}_merged.json"
    save_json(unified, merged_path)
    paths["merged_extraction"] = str(merged_path)

    # 4. Unresolved identifiers
    unresolved_path = intermediate_dir / f"{paper_id}_unresolved_ids.json"
    save_json(id_dict.get("unresolved", {}), unresolved_path)
    paths["unresolved_ids"] = str(unresolved_path)

    # 5. Completeness reports
    comp_path = intermediate_dir / f"{paper_id}_completeness.json"
    save_json(completeness_reports, comp_path)
    paths["completeness_reports"] = str(comp_path)

    # 6. CSV export — routed to solution or solid template based on phase
    csv_path, unresolved_rows = export_to_csv(
        paper_id, final_schemes, output_dir,
        doi=doi, si_data=si_data or {}, id_dict=id_dict,
    )
    if csv_path:
        paths["csv_export"] = str(csv_path)

    # 7. Unresolved fields CSV
    if unresolved_rows:
        unresolved_csv_path = output_dir / f"{paper_id}_unresolved_fields.csv"
        _write_unresolved_csv(unresolved_csv_path, unresolved_rows)
        paths["unresolved_fields_csv"] = str(unresolved_csv_path)
        logger.info(
            f"[post_process_and_save] {len(unresolved_rows)} unresolved field(s) → "
            f"{unresolved_csv_path}"
        )

    # 8. Extraction log (module-level audit trail)
    extraction_log = _build_extraction_log(
        paper_id=paper_id,
        id_dict=id_dict,
        scheme_extractions=scheme_extractions,
        completeness_reports=completeness_reports,
        si_data=si_data or {},
        final_schemes=final_schemes,
    )
    log_path = output_dir / f"{paper_id}_extraction_log.json"
    save_json(extraction_log, log_path)
    paths["extraction_log"] = str(log_path)

    logger.info(f"Saved all outputs for {paper_id}:")
    for name, path in paths.items():
        logger.info(f"  {name}: {path}")

    return paths


# ── Internal helpers (new path) ───────────────────────────────────────────────

def _postprocess_with_llm(fill_result: dict, prompt_text: str, figure_id: str) -> dict:
    """One GPT-4o call for Module 7 standardization + provenance."""
    if not settings.OPENAI_API_KEY:
        return _postprocess_rule_based(fill_result, figure_id)

    user_content = (
        f"Filled reaction extraction JSON:\n"
        f"{json.dumps(fill_result, indent=2)}"
    )

    try:
        from openai import OpenAI
        client = OpenAI(api_key=settings.OPENAI_API_KEY)
        resp = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": prompt_text},
                {"role": "user",   "content": user_content},
            ],
            response_format={"type": "json_object"},
            max_tokens=4096,
        )
        raw    = resp.choices[0].message.content
        result = json.loads(raw)
        logger.info(f"[post_process_and_save] Post-processing complete for {figure_id}")
        return result

    except Exception as exc:
        logger.warning(f"[post_process_and_save] LLM call failed for {figure_id}: {exc}")
        return _postprocess_rule_based(fill_result, figure_id)


def _postprocess_rule_based(fill_result: dict, figure_id: str) -> dict:
    """
    Rule-based post-processing: standardize reaction_type and phase labels,
    set null for missing values, build provenance summary.
    """
    VALID_REACTION_TYPES = {
        "glycosylation", "protection", "deprotection",
        "global_deprotection", "other", "unknown",
    }
    VALID_PHASES = {"solution", "solid", "mixed", "unknown"}

    filled = fill_result.get("filled_output", {})
    steps  = filled.get("reaction_steps", [])

    # Standardize phase.
    phase = filled.get("phase", "unknown")
    if phase not in VALID_PHASES:
        phase = "unknown"

    clean_steps = []
    for step in steps:
        step = dict(step)
        rt = step.get("reaction_type", "unknown")
        if rt not in VALID_REACTION_TYPES:
            step["reaction_type"] = "unknown"
        clean_steps.append(step)

    final_output = {
        "phase":             phase,
        "scheme_steps_count": len(clean_steps),
        "reaction_steps":    clean_steps,
    }

    # Simple provenance summary.
    has_conditions = any(
        s.get("solution_conditions") or s.get("solid_conditions")
        for s in clean_steps
        if s.get("reaction_type") == "glycosylation"
    )
    provenance_summary = {
        "figure_used":              True,
        "main_text_used":           False,
        "SI_used":                  False,
        "compound_dictionary_used": False,
        "manual_check_required":    bool(fill_result.get("unfilled_fields")),
    }

    unresolved = fill_result.get("unfilled_fields", [])

    return {
        "final_output":     final_output,
        "provenance_summary": provenance_summary,
        "unresolved_items": unresolved,
    }


def _load_prompt() -> str:
    try:
        md    = _PROMPT_PATH.read_text(encoding="utf-8")
        start = md.find("```text\n")
        if start == -1:
            start = md.find("```\n")
        end = md.find("\n```", start + 4)
        if start != -1 and end != -1:
            return md[start + md[start:].find("\n") + 1 : end].strip()
        return md.strip()
    except Exception as exc:
        logger.warning(f"[post_process_and_save] Could not load prompt: {exc}")
        return ""


# ── CSV export ───────────────────────────────────────────────────────────────

# Column headers matching the two ground-truth templates exactly.
# "Confidence" column added: high (SI-verified) / medium (figure) / low (filled/guessed).
_SOLUTION_COLUMNS = [
    "DOI", "Donor_ID", "Donor_Name", "Donor_SMILES",
    "Donor_Mass_mg", "Donor_mmol",
    "Acceptor_ID", "Acceptor_Name", "Acceptor_SMILES",
    "Acceptor_Mass_mg", "Acceptor_mmol", "Equiv.",
    "Activator_1", "Activator_1_Mass_mg", "Activator_1_mmol",
    "Activator_2", "Activator_2_Volume_uL", "Activator_2_mmol",
    "Solvent_Name", "Solvent_Volume_mL",
    "Temperature_1_initial", "Temperature_final_Celsius",
    "Reaction_Time_min",
    "Product_ID", "Product_Name", "Product_Mass_mg",
    "a:b_ratio", "Yield(%)", "Step", "Comments", "Confidence",
]

_SOLID_COLUMNS = [
    "DOI", "Donor_ID", "Donor_Name", "Donor_SMILES",
    "Acceptor_ID", "Deprotected_Group_Type",
    "Donor_mmol", "Equiv.", "Resin_mass_mg",
    "Activator_Name", "Activator_1_volume_mL", "Activator_1_mmol",
    "Solvent_Name", "Solvent_Volume_mL",
    "T1_Celsius", "t1_min", "T2_Celcius", "t2_min",
    "Product_mass_mg", "Product_Name",
    "Yield(%)", "a:b_ratio", "Step", "Comments", "Confidence",
]

# Fields that should display "NR" (not resolved) rather than blank
# when no value was extracted.  ID/DOI/Step fields are excluded.
_NR_FIELDS_SOLUTION = {
    "Donor_Name", "Donor_SMILES", "Donor_Mass_mg", "Donor_mmol",
    "Acceptor_Name", "Acceptor_SMILES", "Acceptor_Mass_mg", "Acceptor_mmol",
    "Equiv.", "Activator_1", "Activator_1_Mass_mg", "Activator_1_mmol",
    "Activator_2", "Activator_2_Volume_uL", "Activator_2_mmol",
    "Solvent_Name", "Solvent_Volume_mL",
    "Temperature_1_initial", "Temperature_final_Celsius",
    "Reaction_Time_min", "Product_Name", "Product_Mass_mg",
    "a:b_ratio", "Yield(%)", "Comments",
}

_NR_FIELDS_SOLID = {
    "Donor_Name", "Donor_SMILES",
    "Deprotected_Group_Type", "Donor_mmol", "Equiv.", "Resin_mass_mg",
    "Activator_Name", "Activator_1_volume_mL", "Activator_1_mmol",
    "Solvent_Name", "Solvent_Volume_mL",
    "T1_Celsius", "t1_min", "T2_Celcius", "t2_min",
    "Product_mass_mg", "Product_Name",
    "Yield(%)", "a:b_ratio", "Comments",
}


def _nr(val, field: str, nr_fields: set) -> str:
    """
    Return the value as a string if present; otherwise return "NR" for
    fields that are expected to have data, or "" for ID/structural fields.
    """
    if val is None or str(val).strip() in ("", "null", "None"):
        return "NR" if field in nr_fields else ""
    return str(val)


def _row_confidence(si: dict, c: dict) -> str:
    """
    Derive a per-row confidence label based on which source filled the data.

    high   — SI experimental paragraph found and contributed masses/mmol
    medium — Conditions came from the figure scheme image
    low    — Conditions were filled/guessed by Module 6
    """
    # SI provided quantitative data → high confidence
    si_quantitative = [
        si.get("donor_mass_mg"), si.get("donor_mmol"),
        si.get("acceptor_mass_mg"), si.get("acceptor_mmol"),
        si.get("product_mass_mg"),
    ]
    if any(v not in (None, "", "null") for v in si_quantitative):
        return "high"
    # Figure conditions present → medium
    fig_quantitative = [
        c.get("yield_percent"), c.get("temperature_initial_celsius"),
        c.get("activator_1_name"), c.get("solvent_name"),
    ]
    if any(v not in (None, "", "null") for v in fig_quantitative):
        return "medium"
    return "low"


def export_to_csv(
    paper_id: str,
    final_schemes: List[dict],
    output_dir: Path,
    doi: str = "",
    si_data: Optional[dict] = None,
    id_dict: Optional[dict] = None,
) -> tuple:
    """
    Export all glycosylation steps from final_schemes to a CSV file.

    Routes to the solution-phase or solid-phase column template based on
    the phase field in each scheme's final_output.  If a paper contains both
    phases (rare), two separate files are written.

    Parameters
    ----------
    paper_id      : Used to name the output file.
    final_schemes : List of post-processed scheme dicts from post_process_and_save().
    output_dir    : Where to write the CSV.
    doi           : Optional DOI string to fill the DOI column.

    Returns
    -------
    (last_csv_path_or_None, unresolved_rows_list)
    where unresolved_rows_list contains dicts describing each NR field.
    """
    solution_rows: List[dict] = []
    solid_rows:    List[dict] = []
    unresolved_rows: List[dict] = []   # collected across all schemes/steps
    si_data = si_data or {}

    for scheme in final_schemes:
        final = scheme.get("final_output", {})
        phase = final.get("phase", "unknown").lower()
        steps = final.get("reaction_steps", [])

        for step in steps:
            if step.get("reaction_type") != "glycosylation":
                continue

            step_num = step.get("step", "")

            # ── Helper to safely get a nested compound field ──────────────────
            def _cid(role):
                v = step.get(role)
                return v.get("compound_id", "") if isinstance(v, dict) else (v or "")

            def _cname(role):
                v = step.get(role)
                return v.get("compound_name", "") if isinstance(v, dict) else ""

            donor_id   = _cid("donor")
            donor_name = _cname("donor")
            acc_id     = _cid("acceptor")
            acc_name   = _cname("acceptor")
            prod_id    = _cid("product")
            prod_name  = _cname("product")

            # ── Fall back to id_dict for any missing names ────────────────────
            resolved = (id_dict or {}).get("resolved", {})
            def _lookup_name(cid, name):
                if name:
                    return name
                entry = resolved.get(str(cid)) or resolved.get(cid)
                if entry:
                    return entry.get("compound_name") or entry.get("possible_name") or ""
                return ""
            donor_name = _lookup_name(donor_id, donor_name)
            acc_name   = _lookup_name(acc_id,   acc_name)
            prod_name  = _lookup_name(prod_id,  prod_name)

            # ── SI data for this product (fills masses, mmol, SMILES) ─────────
            # si_data is keyed by product_id → experimental dict
            si = si_data.get(str(prod_id), {}) or {}

            figure_id = scheme.get("figure_id", "unknown")

            if phase == "solid":
                c = step.get("solid_conditions") or {}
                if not c:
                    # Fall back: Module 7 may have used a flat "conditions" key
                    flat = step.get("conditions") or {}
                    c = {
                        "activator_name":        flat.get("promoter_or_activator", ""),
                        "solvent_name":          flat.get("solvent", ""),
                        "T1_celsius":            flat.get("temperature", ""),
                        "t1_min":                flat.get("time", ""),
                        "yield_percent":         flat.get("yield", ""),
                        "a_b_ratio":             flat.get("stereoselectivity", ""),
                    }

                def _vs(si_key, c_key=None):
                    si_val = si.get(si_key)
                    if si_val not in (None, "", "null"):
                        return si_val
                    return c.get(c_key or si_key, "") or ""

                confidence = _row_confidence(si, c)
                row = {
                    "DOI":                   doi,
                    "Donor_ID":              si.get("donor_id") or donor_id,
                    "Donor_Name":            _nr(donor_name, "Donor_Name", _NR_FIELDS_SOLID),
                    "Donor_SMILES":          _nr(_vs("donor_smiles"), "Donor_SMILES", _NR_FIELDS_SOLID),
                    "Acceptor_ID":           si.get("acceptor_id") or acc_id,
                    "Deprotected_Group_Type": _nr(_vs("deprotected_group_type"), "Deprotected_Group_Type", _NR_FIELDS_SOLID),
                    "Donor_mmol":            _nr(_vs("donor_mmol"), "Donor_mmol", _NR_FIELDS_SOLID),
                    "Equiv.":                _nr(_vs("equivalents"), "Equiv.", _NR_FIELDS_SOLID),
                    "Resin_mass_mg":         _nr(_vs("resin_mass_mg"), "Resin_mass_mg", _NR_FIELDS_SOLID),
                    "Activator_Name":        _nr(_vs("activator_1_name", "activator_name"), "Activator_Name", _NR_FIELDS_SOLID),
                    "Activator_1_volume_mL": _nr(_vs("activator_1_volume_mL", "activator_volume_mL"), "Activator_1_volume_mL", _NR_FIELDS_SOLID),
                    "Activator_1_mmol":      _nr(_vs("activator_1_mmol", "activator_mmol"), "Activator_1_mmol", _NR_FIELDS_SOLID),
                    "Solvent_Name":          _nr(_vs("solvent_name"), "Solvent_Name", _NR_FIELDS_SOLID),
                    "Solvent_Volume_mL":     _nr(_vs("solvent_volume_mL"), "Solvent_Volume_mL", _NR_FIELDS_SOLID),
                    "T1_Celsius":            _nr(_vs("temperature_initial_celsius", "T1_celsius"), "T1_Celsius", _NR_FIELDS_SOLID),
                    "t1_min":                _nr(_vs("t1_min"), "t1_min", _NR_FIELDS_SOLID),
                    "T2_Celcius":            _nr(_vs("temperature_final_celsius", "T2_celsius"), "T2_Celcius", _NR_FIELDS_SOLID),
                    "t2_min":                _nr(_vs("t2_min"), "t2_min", _NR_FIELDS_SOLID),
                    "Product_mass_mg":       _nr(_vs("product_mass_mg"), "Product_mass_mg", _NR_FIELDS_SOLID),
                    "Product_Name":          _nr(prod_name, "Product_Name", _NR_FIELDS_SOLID),
                    "Yield(%)":              _nr(_vs("yield_percent"), "Yield(%)", _NR_FIELDS_SOLID),
                    "a:b_ratio":             _nr(_vs("a_b_ratio"), "a:b_ratio", _NR_FIELDS_SOLID),
                    "Step":                  step_num,
                    "Comments":              _nr(_vs("comments"), "Comments", _NR_FIELDS_SOLID),
                    "Confidence":            confidence,
                }
                solid_rows.append(row)
                _collect_unresolved(unresolved_rows, row, _NR_FIELDS_SOLID,
                                    paper_id, figure_id, step_num)
            else:
                c = step.get("solution_conditions") or {}
                if not c:
                    # Fall back: Module 7 may have used a flat "conditions" key
                    flat = step.get("conditions") or {}
                    c = {
                        "activator_1_name":              flat.get("promoter_or_activator", ""),
                        "solvent_name":                  flat.get("solvent", ""),
                        "temperature_initial_celsius":   flat.get("temperature", ""),
                        "temperature_final_celsius":     flat.get("temperature", ""),
                        "reaction_time_min":             flat.get("time", ""),
                        "yield_percent":                 flat.get("yield", ""),
                        "a_b_ratio":                     flat.get("stereoselectivity", ""),
                    }

                # ── Helper: SI value > figure value ───────────────────────────
                def _v(si_key, c_key=None):
                    """Return SI value if present, else figure-extracted value."""
                    si_val = si.get(si_key)
                    if si_val not in (None, "", "null"):
                        return si_val
                    return c.get(c_key or si_key, "") or ""

                confidence = _row_confidence(si, c)
                row = {
                    "DOI":                    doi,
                    "Donor_ID":               si.get("donor_id") or donor_id,
                    "Donor_Name":             _nr(donor_name, "Donor_Name", _NR_FIELDS_SOLUTION),
                    "Donor_SMILES":           _nr(_v("donor_smiles"), "Donor_SMILES", _NR_FIELDS_SOLUTION),
                    "Donor_Mass_mg":          _nr(_v("donor_mass_mg"), "Donor_Mass_mg", _NR_FIELDS_SOLUTION),
                    "Donor_mmol":             _nr(_v("donor_mmol"), "Donor_mmol", _NR_FIELDS_SOLUTION),
                    "Acceptor_ID":            si.get("acceptor_id") or acc_id,
                    "Acceptor_Name":          _nr(acc_name, "Acceptor_Name", _NR_FIELDS_SOLUTION),
                    "Acceptor_SMILES":        _nr(_v("acceptor_smiles"), "Acceptor_SMILES", _NR_FIELDS_SOLUTION),
                    "Acceptor_Mass_mg":       _nr(_v("acceptor_mass_mg"), "Acceptor_Mass_mg", _NR_FIELDS_SOLUTION),
                    "Acceptor_mmol":          _nr(_v("acceptor_mmol"), "Acceptor_mmol", _NR_FIELDS_SOLUTION),
                    "Equiv.":                 _nr(_v("equivalents"), "Equiv.", _NR_FIELDS_SOLUTION),
                    "Activator_1":            _nr(_v("activator_1_name"), "Activator_1", _NR_FIELDS_SOLUTION),
                    "Activator_1_Mass_mg":    _nr(_v("activator_1_mass_mg"), "Activator_1_Mass_mg", _NR_FIELDS_SOLUTION),
                    "Activator_1_mmol":       _nr(_v("activator_1_mmol"), "Activator_1_mmol", _NR_FIELDS_SOLUTION),
                    "Activator_2":            _nr(_v("activator_2_name"), "Activator_2", _NR_FIELDS_SOLUTION),
                    "Activator_2_Volume_uL":  _nr(_v("activator_2_volume_uL"), "Activator_2_Volume_uL", _NR_FIELDS_SOLUTION),
                    "Activator_2_mmol":       _nr(_v("activator_2_mmol"), "Activator_2_mmol", _NR_FIELDS_SOLUTION),
                    "Solvent_Name":           _nr(_v("solvent_name"), "Solvent_Name", _NR_FIELDS_SOLUTION),
                    "Solvent_Volume_mL":      _nr(_v("solvent_volume_mL"), "Solvent_Volume_mL", _NR_FIELDS_SOLUTION),
                    "Temperature_1_initial":  _nr(_v("temperature_initial_celsius"), "Temperature_1_initial", _NR_FIELDS_SOLUTION),
                    "Temperature_final_Celsius": _nr(_v("temperature_final_celsius"), "Temperature_final_Celsius", _NR_FIELDS_SOLUTION),
                    "Reaction_Time_min":      _nr(_v("reaction_time_min"), "Reaction_Time_min", _NR_FIELDS_SOLUTION),
                    "Product_ID":             prod_id,
                    "Product_Name":           _nr(prod_name, "Product_Name", _NR_FIELDS_SOLUTION),
                    "Product_Mass_mg":        _nr(_v("product_mass_mg"), "Product_Mass_mg", _NR_FIELDS_SOLUTION),
                    "a:b_ratio":              _nr(_v("a_b_ratio"), "a:b_ratio", _NR_FIELDS_SOLUTION),
                    "Yield(%)":               _nr(_v("yield_percent"), "Yield(%)", _NR_FIELDS_SOLUTION),
                    "Step":                   step_num,
                    "Comments":               _nr(_v("comments"), "Comments", _NR_FIELDS_SOLUTION),
                    "Confidence":             confidence,
                }
                solution_rows.append(row)
                _collect_unresolved(unresolved_rows, row, _NR_FIELDS_SOLUTION,
                                    paper_id, figure_id, step_num)

    if not solution_rows and not solid_rows:
        logger.warning(f"[export_to_csv] No glycosylation steps found for {paper_id}")
        return None, unresolved_rows

    last_path = None

    if solution_rows:
        path = output_dir / f"{paper_id}_solution.csv"
        _write_csv(path, _SOLUTION_COLUMNS, solution_rows)
        logger.info(f"[export_to_csv] Solution CSV → {path} ({len(solution_rows)} row(s))")
        last_path = path

    if solid_rows:
        path = output_dir / f"{paper_id}_solid.csv"
        _write_csv(path, _SOLID_COLUMNS, solid_rows)
        logger.info(f"[export_to_csv] Solid CSV → {path} ({len(solid_rows)} row(s))")
        last_path = path

    # Always write the Excel file with both tabs
    excel_path = output_dir / "test.xlsx"
    _write_excel(excel_path, _SOLUTION_COLUMNS, solution_rows, _SOLID_COLUMNS, solid_rows)
    logger.info(
        f"[export_to_csv] Excel → {excel_path} "
        f"(solution: {len(solution_rows)} row(s), solid: {len(solid_rows)} row(s))"
    )

    nr_count = len(unresolved_rows)
    if nr_count:
        logger.info(f"[export_to_csv] {nr_count} NR field(s) flagged for unresolved_fields.csv")

    return last_path, unresolved_rows


def _collect_unresolved(
    unresolved_rows: List[dict],
    row: dict,
    nr_fields: set,
    paper_id: str,
    figure_id: str,
    step_num,
) -> None:
    """Append one record per NR field in *row* to *unresolved_rows*."""
    for field in nr_fields:
        if row.get(field) == "NR":
            unresolved_rows.append({
                "paper_id":  paper_id,
                "figure_id": figure_id,
                "step":      step_num,
                "field":     field,
                "reason":    "not found in figure, SI, or fill pass",
            })


def _write_unresolved_csv(path: Path, rows: List[dict]) -> None:
    """Write the unresolved-fields audit CSV."""
    columns = ["paper_id", "figure_id", "step", "field", "reason"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _build_extraction_log(
    paper_id: str,
    id_dict: dict,
    scheme_extractions: list,
    completeness_reports: list,
    si_data: dict,
    final_schemes: list,
) -> dict:
    """
    Build a structured extraction log (audit trail) for this paper.

    Records what each pipeline module produced: counts, quality metrics,
    and key warnings.  Saved as {paper_id}_extraction_log.json.
    """
    import datetime

    resolved   = id_dict.get("resolved", {})
    unresolved = id_dict.get("unresolved", {})

    # Module 3 stats
    m3_figures = len(scheme_extractions)
    m3_steps   = sum(
        len(s.get("step_analysis", {}).get("reaction_steps", []))
        for s in scheme_extractions
    )
    m3_glyco = sum(
        sum(
            1 for step in s.get("step_analysis", {}).get("reaction_steps", [])
            if step.get("reaction_type") == "glycosylation"
        )
        for s in scheme_extractions
    )

    # Module 5 completeness
    complete_count = sum(1 for r in completeness_reports if r.get("is_complete"))

    # Module 2.04 SI stats
    si_found = [cid for cid, d in si_data.items()
                if isinstance(d, dict) and any(v is not None for v in d.values())]

    # Module 7 final output
    final_steps = sum(
        len(s.get("final_output", {}).get("reaction_steps", []))
        for s in final_schemes
    )
    nr_count = sum(
        sum(1 for v in s.get("final_output", {}).get("reaction_steps", []))
        for s in final_schemes
    )

    return {
        "paper_id":   paper_id,
        "generated":  datetime.datetime.utcnow().isoformat() + "Z",
        "modules": {
            "module_2_02_identifier_dictionary": {
                "resolved_count":   len(resolved),
                "unresolved_count": len(unresolved),
                "resolved_ids":     sorted(resolved.keys()),
                "unresolved_ids":   sorted(unresolved.keys()),
            },
            "module_3_figure_extraction": {
                "figures_processed":    m3_figures,
                "total_steps_found":    m3_steps,
                "glycosylation_steps":  m3_glyco,
            },
            "module_5_completeness": {
                "schemes_checked":   len(completeness_reports),
                "schemes_complete":  complete_count,
                "schemes_incomplete": len(completeness_reports) - complete_count,
            },
            "module_2_04_si_extraction": {
                "products_queried": len(si_data),
                "products_found":   len(si_found),
                "found_ids":        sorted(si_found),
            },
            "module_7_postprocessing": {
                "final_schemes": len(final_schemes),
                "final_steps":   final_steps,
            },
        },
    }


def _write_csv(path: Path, columns: List[str], rows: List[dict]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _write_excel(
    path: Path,
    solution_cols: List[str],
    solution_rows: List[dict],
    solid_cols: List[str],
    solid_rows: List[dict],
) -> None:
    """Write test.xlsx with a Solution tab and a Solid tab."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("[export_to_csv] openpyxl not installed — skipping Excel export. Run: pip install openpyxl")
        return

    wb = Workbook()

    # ── Solution tab ──────────────────────────────────────────────────────────
    ws_sol = wb.active
    ws_sol.title = "Solution"
    _fill_sheet(ws_sol, solution_cols, solution_rows)

    # ── Solid tab ─────────────────────────────────────────────────────────────
    ws_sol2 = wb.create_sheet(title="Solid")
    _fill_sheet(ws_sol2, solid_cols, solid_rows)

    wb.save(path)


def _safe_cell_value(v) -> str:
    """Strip XML-illegal control characters so openpyxl never raises IllegalCharacterError."""
    import re
    if v is None:
        return ""
    s = str(v)
    # Remove chars that are illegal in XML 1.0 (openpyxl uses XML internally)
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", s)


def _fill_sheet(ws, columns: List[str], rows: List[dict]) -> None:
    """Write headers + rows into a worksheet with basic formatting."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
        HEADER_FONT  = Font(bold=True, color="FFFFFF")
        HEADER_ALIGN = Alignment(horizontal="center", wrap_text=True)

        # Write header row
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill  = HEADER_FILL
            cell.font  = HEADER_FONT
            cell.alignment = HEADER_ALIGN

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx,
                        value=_safe_cell_value(row_data.get(col_name, "")))

        # Auto-fit column widths (approximate)
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = max(
                len(str(col_name)),
                *(len(str(r.get(col_name, "") or "")) for r in rows) if rows else [0],
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        # Freeze top row
        ws.freeze_panes = "A2"

    except Exception as exc:
        # If styling fails just write plain data
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx,
                        value=_safe_cell_value(row_data.get(col_name, "")))


# ── Legacy path (ReactionRecord) ─────────────────────────────────────────────

def save_outputs(
    record: ReactionRecord,
    unified: dict,
    id_dict: dict,
    output_dir: Optional[Path] = None,
    intermediate_dir: Optional[Path] = None,
) -> dict:
    """
    Legacy: save flat ReactionRecord outputs.
    Kept for backward compatibility.
    """
    output_dir       = Path(output_dir       or settings.OUTPUT_DIR)
    intermediate_dir = Path(intermediate_dir or settings.INTERMEDIATE_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    intermediate_dir.mkdir(parents=True, exist_ok=True)
    paper_id = record.paper_id
    paths    = {}

    final_path = output_dir / f"{paper_id}_final.json"
    save_json(record.to_dict(), final_path)
    paths["final_record"] = str(final_path)

    merged_path = intermediate_dir / f"{paper_id}_merged.json"
    save_json(unified, merged_path)
    paths["merged_extraction"] = str(merged_path)

    unresolved_path = intermediate_dir / f"{paper_id}_unresolved_ids.json"
    save_json(id_dict.get("unresolved", {}), unresolved_path)
    paths["unresolved_ids"] = str(unresolved_path)

    report = {
        "paper_id":           paper_id,
        "completeness_score": record.completeness_score,
        "unresolved_fields":  record.unresolved_fields,
        "provenance":         record.provenance,
    }
    report_path = intermediate_dir / f"{paper_id}_report.json"
    save_json(report, report_path)
    paths["report"] = str(report_path)

    logger.info(f"Saved outputs for {paper_id}:")
    for name, path in paths.items():
        logger.info(f"  {name}: {path}")

    return paths
